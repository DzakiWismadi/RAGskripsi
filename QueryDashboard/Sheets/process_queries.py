import openpyxl
import json

EXCEL_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx'
JSON_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\ragV1\data\iso_controls.json'
SHEET_NAME = 'Explicit Short'
MAX_ROWS = 252  # Process rows 2-252 (Q1-Q250)

def load_iso_controls(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return {item['control_id']: item for item in data}

def get_title(control_id, iso_lookup):
    if control_id is None or control_id == '':
        return None
    control = iso_lookup.get(control_id)
    return control['title'] if control else None

def synthesize_query(gt1_title, gt2_title, gt3_title):
    titles = [gt1_title, gt2_title, gt3_title]
    valid_titles = [t for t in titles if t is not None]

    if not valid_titles:
        return None

    # Build explanation reasoning (Column F)
    explanation_parts = []
    if gt1_title:
        explanation_parts.append(f"GT1: {gt1_title}")
    if gt2_title:
        explanation_parts.append(f"GT2: {gt2_title}")
    if gt3_title:
        explanation_parts.append(f"GT3: {gt3_title}")
    explanation = " | ".join(explanation_parts)

    # Synthesize query (Column B) - describe a scenario based on control titles
    query_parts = []

    for title in valid_titles:
        # Convert each title into a sentence about a gap/issue
        # Common patterns in Indonesian security contexts
        if 'perlindungan' in title.lower() or 'proteksi' in title.lower():
            query_parts.append(f"Sistem belum memiliki {title.lower()} yang memadai")
        elif 'pemantauan' in title.lower() or 'monitoring' in title.lower():
            query_parts.append(f"Tidak terdapat {title.lower()} yang konsisten untuk mendeteksi ancaman")
        elif 'kebijakan' in title.lower():
            query_parts.append(f"{title} belum diterapkan secara efektif di seluruh organisasi")
        elif 'persyaratan' in title.lower():
            query_parts.append(f"{title} belum terdefinisi dengan jelas")
        elif 'autentikasi' in title.lower():
            query_parts.append(f"Mekanisme {title.lower()} belum cukup kuat")
        elif 'penggunaan' in title.lower() or 'pengelolaan' in title.lower() or 'manajemen' in title.lower():
            query_parts.append(f"{title} belum dilakukan sesuai standar")
        elif 'akses' in title.lower():
            query_parts.append(f"{title} belum diatur dengan tepat")
        elif 'keamanan' in title.lower():
            query_parts.append(f"{title} belum memenuhi standar yang ditetapkan")
        elif 'inventarisasi' in title.lower() or 'inventaris' in title.lower():
            query_parts.append(f"{title} aset belum dilakukan secara menyeluruh")
        elif 'pengembalian' in title.lower():
            query_parts.append(f"Prosedur {title.lower()} belum dijalankan dengan baik")
        elif 'konfigurasi' in title.lower():
            query_parts.append(f"{title} belum dikelola secara konsisten")
        elif 'kerentanan' in title.lower():
            query_parts.append(f"{title} belum ditangani secara berkala")
        elif 'pengujian' in title.lower():
            query_parts.append(f"{title} belum dilakukan secara rutin")
        elif 'kontrol' in title.lower():
            query_parts.append(f"{title} belum diterapkan secara optimal")
        else:
            query_parts.append(f"{title} belum sesuai dengan standar keamanan")

    # Combine into a coherent query
    if len(query_parts) == 1:
        query = query_parts[0] + ", sehingga berpotensi menyebabkan kerentanan keamanan."
    elif len(query_parts) == 2:
        query = query_parts[0] + ". Selain itu, " + query_parts[1].lower() + "."
    else:
        query = query_parts[0] + ". " + query_parts[1].lower() + ". " + query_parts[2].lower() + "."

    return explanation, query

def main():
    print("Loading ISO controls...")
    iso_lookup = load_iso_controls(JSON_PATH)
    print(f"Loaded {len(iso_lookup)} controls")

    print("Opening Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print(f"\nProcessing rows 2 to {MAX_ROWS - 1} (first {MAX_ROWS - 2} queries)...")

    for row in range(2, MAX_ROWS):
        query_id = ws.cell(row=row, column=1).value

        if not query_id:
            print(f"Row {row}: No query_id, skipping")
            continue

        # Read ground truth controls (Columns C, D, E)
        gt1 = ws.cell(row=row, column=3).value
        gt2 = ws.cell(row=row, column=4).value
        gt3 = ws.cell(row=row, column=5).value

        # Lookup titles from JSON
        gt1_title = get_title(gt1, iso_lookup)
        gt2_title = get_title(gt2, iso_lookup)
        gt3_title = get_title(gt3, iso_lookup)

        print(f"\n{query_id}: GT1={gt1}, GT2={gt2}, GT3={gt3}")
        print(f"  Titles: {gt1_title} | {gt2_title} | {gt3_title}")

        # Synthesize explanation (Column F) and query (Column B)
        explanation, query = synthesize_query(gt1_title, gt2_title, gt3_title)

        print(f"  Explanation: {explanation}")
        print(f"  Query: {query}")

        # Write to Excel
        ws.cell(row=row, column=6).value = explanation  # Column F
        ws.cell(row=row, column=2).value = query  # Column B

    print("\nSaving workbook...")
    wb.save(EXCEL_PATH)
    print("Done!")

if __name__ == '__main__':
    main()
