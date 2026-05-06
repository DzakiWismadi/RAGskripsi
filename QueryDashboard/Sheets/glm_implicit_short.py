"""
glm_implicit_short.py
---------------------
Reads 'Implicit Short' tab from Query_Source_Counter.xlsx (Q1-Q250),
looks up ISO control titles from iso_controls.json,
writes titles to column F, then calls ZhipuAI GLM-4.7 to generate
an implicit Indonesian query for column B.

Usage:
    python glm_implicit_short.py

Configure the three paths at the top of this file before running.
Requires:  pip install openpyxl requests
"""

import json
import time
import openpyxl
import requests

XLSX_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx"
JSON_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\ragV1\data\iso_controls.json"
SHEET_NAME = "Implicit Short"

START_ROW = 2
END_ROW   = 251

API_KEY = "u#chk8b0ydnzou8udfvkljjuv0wjx5x00j6hb4xks2c9"
BASE_URL = "https://zai.izcy.tech/v1/chat/completions"
MODEL = "glm-4.7"

API_DELAY = 1.0

SYSTEM_PROMPT = """Kamu adalah generator query untuk dataset evaluasi RAG (Retrieval-Augmented Generation) berbasis ISO/IEC 27001:2022 di lingkungan perguruan tinggi Indonesia.

Tugasmu: Buat 1-3 kalimat dalam Bahasa Indonesia yang mendeskripsikan suatu skenario/situasi nyata di universitas, TANPA menyebut nama kontrol ISO secara eksplisit.

PANDUAN ENTITAS PENDIDIKAN (gunakan salah satu atau beberapa):
- Direktorat Keuangan
- Direktorat Kemahasiswaan
- Direktorat Akademik
- UPT TIK / Direktorat Sistem Informasi
- Fakultas
- Program Studi
- Biro Akademik
- LPPM
- Perpustakaan

KATEGORI KATA KUNCI AKSI:
A. Sistem & Infrastruktur: cloud, VPS, server, integrasi sistem, jaringan, infrastruktur teknologi
B. Pengelolaan Data: data mahasiswa, data keuangan, rekam akademik, basis data, informasi sensitif
C. Penggunaan Aplikasi: sistem akademik, portal mahasiswa, aplikasi e-learning, platform digital
D. Akses & Identitas: akun institusi, login, kredensial, hak akses, autentikasi pengguna
E. Aktivitas Operasional: pelayanan, koordinasi, administrasi, prosedur operasional, pengelolaan layanan

GAYA PENULISAN:
- Implisit: Jangan sebut nama kontrol ISO, kode kontrol, atau istilah teknis standar secara langsung
- Naturalistik: Gunakan bahasa narasi seperti laporan situasional atau temuan audit internal
- Kontekstual: Ceritakan situasi/kondisi yang terjadi di lingkungan universitas
- Panjang: 1-3 kalimat saja

OUTPUT: Hanya teks query-nya saja, tanpa penjelasan tambahan, tanpa tanda kutip, tanpa nomor."""


def build_user_prompt(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title):
    lines = [
        "Buat sebuah query implisit berdasarkan kontrol ISO berikut:",
        "",
        f"Kontrol 1: {gt1_id} — {gt1_title}",
    ]
    if gt2_id and gt2_id != "No Value":
        lines.append(f"Kontrol 2: {gt2_id} — {gt2_title}")
    if gt3_id and gt3_id != "No Value":
        lines.append(f"Kontrol 3: {gt3_id} — {gt3_title}")
    lines += [
        "",
        "Tulis 1-3 kalimat situasi nyata di universitas Indonesia yang secara IMPLISIT merujuk pada kontrol-kontrol tersebut.",
        "Jangan sebutkan nama kontrol, kode, atau istilah ISO secara langsung.",
    ]
    return "\n".join(lines)


def build_f_value(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title):
    lines = []
    if gt1_id:
        lines.append(f"{gt1_id} — {gt1_title}")
    if gt2_id and gt2_id != "No Value":
        lines.append(f"{gt2_id} — {gt2_title}")
    elif gt2_id == "No Value" or not gt2_id:
        lines.append("GT2: No Value")
    if gt3_id and gt3_id != "No Value":
        lines.append(f"{gt3_id} — {gt3_title}")
    elif gt3_id == "No Value" or not gt3_id:
        lines.append("GT3: No Value")
    return "\n".join(lines)


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        controls = json.load(f)
    lookup = {item["control_id"]: item["title"] for item in controls}

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    if ws["F1"].value is None:
        ws["F1"].value = "Explanation Reasoning"

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json",
    }

    total = END_ROW - START_ROW + 1
    processed = 0
    skipped = 0
    max_retries = 3

    for row in range(START_ROW, END_ROW + 1):
        query_id = ws.cell(row=row, column=1).value
        gt1_raw  = ws.cell(row=row, column=3).value
        gt2_raw  = ws.cell(row=row, column=4).value
        gt3_raw  = ws.cell(row=row, column=5).value

        if not gt1_raw:
            skipped += 1
            print(f"[{row}] {query_id or 'N/A'} -- skipped (no GT1)")
            continue

        existing = ws.cell(row=row, column=2).value
        if existing and "[ERROR" not in str(existing):
            processed += 1
            continue

        gt1_id = str(gt1_raw).strip() if gt1_raw else None
        gt2_id = str(gt2_raw).strip() if gt2_raw else "No Value"
        gt3_id = str(gt3_raw).strip() if gt3_raw else "No Value"

        gt1_title = lookup.get(gt1_id, f"(title not found for {gt1_id})")
        gt2_title = lookup.get(gt2_id, "No Value") if gt2_id != "No Value" else "No Value"
        gt3_title = lookup.get(gt3_id, "No Value") if gt3_id != "No Value" else "No Value"

        f_val = build_f_value(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title)
        ws.cell(row=row, column=6).value = f_val

        user_msg = build_user_prompt(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title)

        generated_query = None
        for attempt in range(1, max_retries + 1):
            try:
                payload = {
                    "model": MODEL,
                    "max_tokens": 400,
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_msg},
                    ],
                }
                resp = requests.post(BASE_URL, json=payload, headers=headers, timeout=60)
                resp.raise_for_status()
                data = resp.json()
                generated_query = data["choices"][0]["message"]["content"].strip()
                break
            except Exception as e:
                print(f"  !! Attempt {attempt}/{max_retries} failed at row {row}: {e}")
                if attempt < max_retries:
                    time.sleep(3)

        if not generated_query:
            generated_query = f"[ERROR: failed after {max_retries} attempts]"

        ws.cell(row=row, column=2).value = generated_query

        processed += 1
        pct = int(processed / total * 100)
        print(f"[{row}] {query_id} -> GT1={gt1_id} | GT2={gt2_id} | GT3={gt3_id}")
        print(f"       F: {f_val.replace(chr(10), ' | ')}")
        print(f"       B: {generated_query[:100]}{'...' if len(generated_query)>100 else ''}")
        print(f"       Progress: {processed}/{total} ({pct}%)")
        print()

        if processed % 10 == 0:
            wb.save(XLSX_PATH)
            print(f"  -- checkpoint saved ({processed} rows done) --\n")

        time.sleep(API_DELAY)

    wb.save(XLSX_PATH)
    print("=" * 60)
    print(f"Done. Processed: {processed} rows | Skipped: {skipped} rows")
    print(f"Saved to: {XLSX_PATH}")


if __name__ == "__main__":
    main()
