import openpyxl
import json
import random

EXCEL_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx'
JSON_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\ragV1\data\iso_controls.json'
SHEET_NAME = 'Implicit Short'
START_ROW = 2
END_ROW = 251

def load_iso_controls(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return {item['control_id']: item for item in data}

def get_control_info(control_id, iso_lookup):
    if control_id is None or control_id == '':
        return None
    return iso_lookup.get(control_id)

INSTITUTIONS = [
    "Direktorat Keuangan",
    "Direktorat Kemahasiswaan",
    "Direktorat Akademik",
    "UPT TIK",
    "Direktorat Sistem Informasi",
    "Fakultas Teknik",
    "Fakultas Ekonomi dan Bisnis",
    "Fakultas Ilmu Sosial dan Politik",
    "Fakultas Kedokteran",
    "Fakultas MIPA",
    "Fakultas Hukum",
    "Program Studi Teknik Informatika",
    "Program Studi Sistem Informasi",
    "Program Studi Manajemen",
    "Biro Akademik",
    "Biro Keuangan",
    "Biro Umum",
    "LPPM",
    "Perpustakaan",
    "Bagian Registrasi",
    "UPT Bahasa",
    "Puslitbang",
    "Biro SDM",
    "Wakil Rektor Bidang Akademik",
    "Wakil Rektor Bidang Keuangan",
    "Kantor Urusan Internasional",
    "Pusat Karir",
    "Lembaga Penjaminan Mutu",
    "UPT Laboratorium Terpadu",
]

SYSTEM_INFRA = [
    "cloud", "VPS", "server", "server utama", "server kampus",
    "jaringan kampus", "jaringan lokal", "jaringan nirkabel",
    "infrastruktur jaringan", "sistem jaringan", "bandwidth",
    "firewall", "proxy", "gateway", "router", "switch",
    "data center", "ruang server", "cloud computing", "virtualisasi",
    "integration", "integrasi sistem", "API", "load balancer",
    "backup server", "disaster recovery site", "storage",
    "database server", "web server", "mail server",
    "sistem operasi", "firmware", "LAN", "WAN",
]

DATA_HANDLING = [
    "data mahasiswa", "data keuangan", "data dosen", "data pegawai",
    "data akademik", "data penelitian", "data nilai", "data transkrip",
    "data alumni", "data pendaftar", "data beasiswa", "data kurikulum",
    "data inventaris", "data kepegawaian", "data pengadaan",
    "data penjaminan mutu", "data perpustakaan", "data laboratorium",
    "data inventaris aset", "data registrasi", "data wisuda",
    "data KKM", "data KRS", "data KTP", "data NPWP",
    "rekapitulasi keuangan", "laporan semester", "rekaman kehadiran",
]

APP_USAGE = [
    "sistem akademik", "portal mahasiswa", "sistem informasi akademik",
    "SIAKAD", "SIMKEU", "SIMPEG", "e-learning", " Learning Management System",
    "sistem perpustakaan digital", "sistem registrasi online",
    "aplikasi pendaftaran", "sistem penjadwalan", "sistem presensi",
    "portal dosen", "portal keuangan", "aplikasi beasiswa",
    "sistem ujian online", "aplikasi survei", "sistem pengaduan",
    "dashboard rektorat", "sistem surat menyurat", "SIMAK",
    "sistem informasi kemahasiswaan", "aplikasi KRS online",
    "sistem informasi penelitian", "portal alumni",
    "aplikasi pelayanan terpadu", "sistem monitoring",
]

ACCESS_IDENTITY = [
    "akun institusi", "login", "kredensial", "password",
    "akun email kampus", "akses VPN", "Single Sign-On",
    "autentikasi dua faktor", "token akses", "hak akses",
    "akun administrator", "akun pengguna", "ID pegawai",
    "akses Wi-Fi kampus", "kartu identitas digital", "sertifikat digital",
    "akses basisdata", "role pengguna", "perizinan akses",
    "kredensial jaringan", "akun mahasiswa", "NIDN", "NIM",
]

OPERATIONAL = [
    "pelayanan", "koordinasi", "administrasi", "pengelolaan",
    "penyelenggaraan", "pelaporan", "pendaftaran", "verifikasi",
    "pemeliharaan", "pengawasan", "penilaian", "audit",
    "pengujian", "pelatihan", "sosialisasi", "evaluasi",
    "penyusunan", "pengembangan", "perencanaan", "implementasi",
    "pemantauan", "pengendalian", "dokumentasi", "revisi",
    "pencatatan", "persuratan", "pengarsipan", "diseminasi",
]

IMPLICIT_PHRASES = [
    "dalam rangka menunjang kelancaran operasional",
    "sebagai bagian dari tata kelola yang berjalan",
    "dalam konteks pengelolaan layanan yang ada",
    "guna mendukung kelangsungan aktivitas sehari-hari",
    "sebagai upaya menjaga kelangsungan fungsi layanan",
    "dalam rangka meningkatkan kualitas pelayanan",
    "untuk mendukung proses yang sedang berlangsung",
    "sebagai bagian dari peningkatan layanan",
    "dalam upaya mengoptimalkan kinerja unit kerja",
    "demi kelancaran proses administratif",
    "sebagai langkah dalam mendukung operasional",
    "dalam kerangka peningkatan kualitas pengelolaan",
    "guna menjamin kelancaran alur kerja",
    "untuk memastikan ketersediaan layanan",
    "sebagai wujud peningkatan efektivitas kerja",
]

CONNECTORS = [
    "Selain itu,",
    "Di samping itu,",
    "Lebih lanjut,",
    "Pada aspek lain,",
    "Dalam konteks yang berkaitan,",
    "Terkait dengan hal tersebut,",
    "Sejalan dengan itu,",
    "Berkaitan dengan aspek tersebut,",
    "Dalam hubungannya dengan upaya tersebut,",
    "Selanjutnya,",
]

TITLE_TO_KEYWORD_MAP = {
    'perlindungan': ['perlindungan', 'pencegahan', 'penangkalan', 'pengamanan'],
    'malware': ['perangkat lunak berbahaya', 'virus', 'malware', 'ancaman siber', 'Anti Virus'],
    'pemantauan': ['pengamatan', 'pemantauan', 'monitoring', 'pengawasan berkala', 'deteksi dini'],
    'keamanan': ['keamanan', 'pengamanan', 'proteksi', 'ketahanan'],
    'informasi': ['informasi', 'data', 'rekaman', 'dokumentasi'],
    'jaringan': ['jaringan', 'konektivitas', 'infrastruktur komunikasi', 'kabel', 'nirkabel'],
    'akses': ['akses', 'konektivitas', 'hak masuk', 'izin penggunaan'],
    'autentikasi': ['autentikasi', 'verifikasi identitas', 'login', 'pembuktian identitas'],
    'enkripsi': ['enkripsi', 'penyandian', 'kriptografi', 'pengkodean'],
    'cadangan': ['cadangan', 'backup', 'salinan data', 'duplikasi'],
    'kerentanan': ['kerentanan', 'kelemahan', 'celah', 'titik rawan'],
    'konfigurasi': ['konfigurasi', 'pengaturan', 'setup', 'parameter'],
    'kebijakan': ['kebijakan', 'aturan', 'pedoman', 'ketentuan'],
    'manajemen': ['manajemen', 'pengelolaan', 'pengaturan', 'koordinasi'],
    'izin': ['izin', 'otorisasi', 'persetujuan', 'kewenangan'],
    'audit': ['audit', 'pemeriksaan', 'inspeksi', 'evaluasi'],
    'log': ['log', 'catatan aktivitas', 'riwayat', 'rekaman'],
    'keamanan fisik': ['keamanan fisik', 'pengamanan ruang', 'kontrol akses fisik'],
    'insiden': ['insiden', 'kejadian', 'gangguan', 'pelanggaran'],
    'privasi': ['privasi', 'kerahasiaan', 'perlindungan data pribadi'],
    'sistem': ['sistem', 'platform', 'aplikasi', 'infrastruktur'],
    'clouD': ['cloud', 'awan', 'layanan cloud', 'penyimpanan awan'],
    'kapasitas': ['kapasitas', 'daya tampung', 'sumber daya', 'skalabilitas'],
    'dokumentasi': ['dokumentasi', 'pendokumentasian', 'pencatatan', 'dokumen'],
    'perangkat': ['perangkat', 'devais', 'hardware', 'perangkat keras'],
    'perangkat lunak': ['perangkat lunak', 'software', 'aplikasi', 'program'],
    'pengguna': ['pengguna', 'user', 'pengguna akhir', 'pegawai'],
    'identitas': ['identitas', 'ID', 'profil', 'kredensial'],
    'kontrol': ['kontrol', 'pengendalian', 'monitoring', 'supervisi'],
    'pemulihan': ['pemulihan', 'recovery', 'pengembalian', 'restorasi'],
    'penyimpanan': ['penyimpanan', 'storage', 'gudang data', 'arsip'],
    'transfer': ['transfer', 'pengiriman', 'pertukaran', 'distribusi'],
    'pengembangan': ['pengembangan', 'development', 'penyempurnaan', 'peningkatan'],
    'pengujian': ['pengujian', 'testing', 'uji coba', 'percobaan'],
    'pelatihan': ['pelatihan', 'training', 'pendidikan', 'sosialisasi'],
    'perubahan': ['perubahan', 'modifikasi', 'penyesuaian', 'revisi'],
    'layanan': ['layanan', 'servis', 'pelayanan', 'fasilitas'],
    'operasional': ['operasional', 'operasi', 'kinerja', 'aktivitas'],
}

def map_title_to_keywords(title):
    title_lower = title.lower()
    keywords = []
    for key, values in TITLE_TO_KEYWORD_MAP.items():
        if key in title_lower:
            keywords.extend(values)
    if not keywords:
        keywords = [title_lower]
    return keywords

def synthesize_implicit_short_query(gt1_title, gt2_title, gt3_title, rng):
    titles = []
    if gt1_title:
        titles.append(('GT1', gt1_title))
    if gt2_title:
        titles.append(('GT2', gt2_title))
    if gt3_title:
        titles.append(('GT3', gt3_title))

    if not titles:
        return None

    institution = rng.choice(INSTITUTIONS)
    system = rng.choice(SYSTEM_INFRA)
    data = rng.choice(DATA_HANDLING)
    app = rng.choice(APP_USAGE)
    access = rng.choice(ACCESS_IDENTITY)
    ops = rng.choice(OPERATIONAL)
    implicit_phrase = rng.choice(IMPLICIT_PHRASES)

    def build_sentence_for_title(rank, title, is_first=True):
        title_lower = title.lower()
        keywords = map_title_to_keywords(title)

        if is_first:
            templates = [
                "{institution} melakukan {ops} {system} yang digunakan untuk {app}. "
                "Penggunaan {kw} pada sistem tersebut berperan dalam {implicit_phrase} "
                "dan membantu menjaga kelangsungan {data}.",

                "{institution} menjalankan {ops} terhadap {app} dalam konteks {data}. "
                "Aktivitas {kw} yang dilakukan secara berkala mendukung kelancaran "
                "proses {access} di lingkungan unit kerja.",

                "Pada {app} yang dikelola oleh {institution}, terdapat aktivitas {ops} "
                "yang mencakup {kw} sebagai bagian dari {implicit_phrase}. "
                "Hal ini berkaitan dengan pengelolaan {data} dan {system}.",

                "{institution} mengelola {system} untuk mendukung {app} "
                "dalam pengelolaan {data}. Aktivitas {kw} dilakukan "
                "{implicit_phrase} dan terintegrasi dengan mekanisme {access}.",

                "Dalam rangka {ops} {app}, {institution} menerapkan "
                "{kw} pada {system} untuk mendukung pengelolaan {data}. "
                "Langkah ini merupakan bagian dari {implicit_phrase}.",
            ]
        else:
            templates = [
                "{connector} terdapat {kw} yang berjalan secara berkala "
                "untuk mendukung {ops} {app} dan pengelolaan {data}.",

                "{connector} aktivitas {kw} juga dilaksanakan pada {system} "
                "sebagai bagian dari {ops} {data} yang dilakukan oleh {institution}.",

                "{connector} proses {kw} terintegrasi dalam {app} "
                "untuk mendukung {implicit_phrase} dan menjaga "
                "ketersediaan {data}.",

                "{connector} {institution} juga melaksanakan {kw} "
                "pada {system} sebagai bagian dari {ops} dan "
                "pengelolaan {access}.",

                "{connector} terdapat upaya {kw} yang dilakukan secara berkala "
                "pada {system} untuk mendukung kelancaran {ops} {app}.",
            ]

        kw = keywords[0] if keywords else title_lower

        if is_first:
            template = rng.choice(templates)
            return template.format(
                institution=institution,
                system=system,
                data=data,
                app=app,
                access=access,
                ops=ops,
                kw=kw,
                implicit_phrase=implicit_phrase,
            )
        else:
            connector = rng.choice(CONNECTORS)
            template = rng.choice(templates)
            return template.format(
                connector=connector,
                kw=kw,
                app=app,
                data=data,
                system=system,
                ops=ops,
                institution=institution,
                access=access,
                implicit_phrase=implicit_phrase,
            )

    sentences = []
    for i, (rank, title) in enumerate(titles):
        is_first = (i == 0)
        sentence = build_sentence_for_title(rank, title, is_first)
        sentences.append(sentence)

    return " ".join(sentences)

def build_explanation(gt1, gt2, gt3, iso_lookup):
    parts = []
    info1 = get_control_info(gt1, iso_lookup)
    if info1:
        parts.append("GT1: " + info1['title'])
    else:
        parts.append("GT1: No Value")
    info2 = get_control_info(gt2, iso_lookup) if gt2 else None
    if info2:
        parts.append("GT2: " + info2['title'])
    else:
        parts.append("GT2: No Value")
    info3 = get_control_info(gt3, iso_lookup) if gt3 else None
    if info3:
        parts.append("GT3: " + info3['title'])
    else:
        parts.append("GT3: No Value")
    return " | ".join(parts)

def main():
    print("Loading ISO controls...")
    iso_lookup = load_iso_controls(JSON_PATH)
    print(f"Loaded {len(iso_lookup)} controls")

    print("Opening Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print(f"\nProcessing rows {START_ROW} to {END_ROW} ({END_ROW - START_ROW + 1} queries)...\n")

    processed = 0
    for row in range(START_ROW, END_ROW + 1):
        query_id = ws.cell(row=row, column=1).value
        gt1 = ws.cell(row=row, column=3).value
        gt2 = ws.cell(row=row, column=4).value
        gt3 = ws.cell(row=row, column=5).value

        if not gt1:
            print(f"Row {row}: No GT1, skipping")
            continue

        explanation = build_explanation(gt1, gt2, gt3, iso_lookup)

        gt1_info = get_control_info(gt1, iso_lookup)
        gt1_title = gt1_info['title'] if gt1_info else None
        gt2_info = get_control_info(gt2, iso_lookup) if gt2 else None
        gt2_title = gt2_info['title'] if gt2_info else None
        gt3_info = get_control_info(gt3, iso_lookup) if gt3 else None
        gt3_title = gt3_info['title'] if gt3_info else None

        rng = random.Random(hash((gt1, gt2, gt3, row)))
        query = synthesize_implicit_short_query(gt1_title, gt2_title, gt3_title, rng)

        ws.cell(row=row, column=6).value = explanation
        ws.cell(row=row, column=2).value = query

        processed += 1
        if processed <= 10 or processed % 50 == 0:
            print(f"{query_id}: GT1={gt1}, GT2={gt2}, GT3={gt3}")
            print(f"  Explanation: {explanation}")
            if query and len(query) > 150:
                print(f"  Query: {query[:150]}...")
            else:
                print(f"  Query: {query}")
            print()

    wb.save(EXCEL_PATH)
    print(f"\nDone! Processed {processed} queries. Saved to {EXCEL_PATH}")

if __name__ == "__main__":
    main()
