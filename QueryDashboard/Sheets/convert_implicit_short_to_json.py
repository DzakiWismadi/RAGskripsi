"""
convert_implicit_short_to_json.py
----------------------------------
Converts the 'Implicit Short' tab from Query_Source_Counter.xlsx
to JSON format in ImpShort.json
"""

import json
import openpyxl

XLSX_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx"
OUTPUT_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\static\ImpShort.json"
SHEET_NAME = "Implicit Short"

START_ROW = 2
END_ROW = 251

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    data = []

    for row in range(START_ROW, END_ROW + 1):
        query_id = ws.cell(row=row, column=1).value
        query = ws.cell(row=row, column=2).value
        gt1 = ws.cell(row=row, column=3).value
        gt2 = ws.cell(row=row, column=4).value
        gt3 = ws.cell(row=row, column=5).value

        if not query_id or not query:
            continue

        ground_truth_ranked = []
        if gt1:
            ground_truth_ranked.append(gt1)
        if gt2:
            ground_truth_ranked.append(gt2)
        if gt3:
            ground_truth_ranked.append(gt3)

        entry = {
            "query_id": query_id,
            "query": query,
            "ground_truth_ranked": ground_truth_ranked
        }
        data.append(entry)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Created {OUTPUT_PATH} with {len(data)} entries")
    print(f"Sample entry (Q1):")
    print(json.dumps(data[0], ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
