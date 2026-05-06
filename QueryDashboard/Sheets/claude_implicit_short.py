"""
generate_implicit_short.py
--------------------------
Reads 'Implicit Short' tab from Query_Source_Counter.xlsx (Q1–Q250),
looks up ISO control titles from iso_controls.json,
writes titles to column F, then calls Claude to generate
an implicit Indonesian query for column B.

Usage:
    python generate_implicit_short.py

Configure the three paths at the top of this file before running.
Requires:  pip install openpyxl anthropic
"""

import json
import time
import openpyxl
import anthropic

# ─────────────────────────────────────────────
# CONFIG  ← edit these three paths
# ─────────────────────────────────────────────
XLSX_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx"
JSON_PATH = r"D:\Hilmi\Coding\MasterFolderSkripsi\RAG\ragV1\data\iso_controls.json"
SHEET_NAME = "Implicit Short"

# Row range (inclusive, 1-indexed; row 1 = header)
START_ROW = 2    # Q1
END_ROW   = 251  # Q250

# Anthropic API — leave as "" to use ANTHROPIC_API_KEY env var
ANTHROPIC_API_KEY = ""
MODEL = ""

# Delay between API calls in seconds (avoid rate limits)
API_DELAY = 1.0
# ─────────────────────────────────────────────

SYSTEM_PROMPT = """Kamu adalah generator query untuk dataset evaluasi RAG (Retrieval-Augmented Generation) berbasis ISO/IEC 27001:2022 di lingkungan perguruan tinggi Indonesia.

Tugasmu: Buat 1–3 kalimat dalam Bahasa Indonesia yang mendeskripsikan suatu skenario/situasi nyata di universitas, TANPA menyebut nama kontrol ISO secara eksplisit.

PANDUAN ENTITAS PENDIDIKAN (gunakan salah satu atau beberapa):
- Direktorat Keuangan
- Direktorat Kemahasiswaan
- Direktorat Akademik
- UPT TIK / Direktorat Sistem Informasi
- Fakultas
- Program Studi
- Biro Akademik
- LPPM
- Perpustakaan

KATEGORI KATA KUNCI AKSI:
A. Sistem & Infrastruktur: cloud, VPS, server, integrasi sistem, jaringan, infrastruktur teknologi
B. Pengelolaan Data: data mahasiswa, data keuangan, rekam akademik, basis data, informasi sensitif
C. Penggunaan Aplikasi: sistem akademik, portal mahasiswa, aplikasi e-learning, platform digital
D. Akses & Identitas: akun institusi, login, kredensial, hak akses, autentikasi pengguna
E. Aktivitas Operasional: pelayanan, koordinasi, administrasi, prosedur operasional, pengelolaan layanan

GAYA PENULISAN:
- Implisit: Jangan sebut nama kontrol ISO, kode kontrol, atau istilah teknis standar secara langsung
- Naturalistik: Gunakan bahasa narasi seperti laporan situasional atau temuan audit internal
- Kontekstual: Ceritakan situasi/kondisi yang terjadi di lingkungan universitas
- Panjang: 1–3 kalimat saja

OUTPUT: Hanya teks query-nya saja, tanpa penjelasan tambahan, tanpa tanda kutip, tanpa nomor."""

def build_user_prompt(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title):
    lines = [
        "Buat sebuah query implisit berdasarkan kontrol ISO berikut:",
        "",
        f"Kontrol 1: {gt1_id} — {gt1_title}",
    ]
    if gt2_id and gt2_id != "No Value":
        lines.append(f"Kontrol 2: {gt2_id} — {gt2_title}")
    if gt3_id and gt3_id != "No Value":
        lines.append(f"Kontrol 3: {gt3_id} — {gt3_title}")
    lines += [
        "",
        "Tulis 1–3 kalimat situasi nyata di universitas Indonesia yang secara IMPLISIT merujuk pada kontrol-kontrol tersebut.",
        "Jangan sebutkan nama kontrol, kode, atau istilah ISO secara langsung.",
    ]
    return "\n".join(lines)


def build_f_value(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title):
    lines = []
    if gt1_id:
        lines.append(f"{gt1_id} — {gt1_title}")
    if gt2_id and gt2_id != "No Value":
        lines.append(f"{gt2_id} — {gt2_title}")
    elif gt2_id == "No Value" or not gt2_id:
        lines.append("GT2: No Value")
    if gt3_id and gt3_id != "No Value":
        lines.append(f"{gt3_id} — {gt3_title}")
    elif gt3_id == "No Value" or not gt3_id:
        lines.append("GT3: No Value")
    return "\n".join(lines)


def main():
    # Load ISO controls lookup
    with open(JSON_PATH, encoding="utf-8") as f:
        controls = json.load(f)
    lookup = {item["control_id"]: item["title"] for item in controls}

    # Load workbook
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    # Set column F header if blank
    if ws["F1"].value is None:
        ws["F1"].value = "Explanation Reasoning"

    # Init Anthropic client
    client_kwargs = {}
    if ANTHROPIC_API_KEY:
        client_kwargs["api_key"] = ANTHROPIC_API_KEY
    client = anthropic.Anthropic(**client_kwargs)

    total = END_ROW - START_ROW + 1
    processed = 0
    skipped = 0

    for row in range(START_ROW, END_ROW + 1):
        query_id = ws.cell(row=row, column=1).value  # col A
        gt1_raw  = ws.cell(row=row, column=3).value  # col C
        gt2_raw  = ws.cell(row=row, column=4).value  # col D
        gt3_raw  = ws.cell(row=row, column=5).value  # col E

        # Skip rows with no ground truth at all
        if not gt1_raw:
            skipped += 1
            print(f"[{row}] {query_id or 'N/A'} — skipped (no GT1)")
            continue

        gt1_id = str(gt1_raw).strip() if gt1_raw else None
        gt2_id = str(gt2_raw).strip() if gt2_raw else "No Value"
        gt3_id = str(gt3_raw).strip() if gt3_raw else "No Value"

        gt1_title = lookup.get(gt1_id, f"(title not found for {gt1_id})")
        gt2_title = lookup.get(gt2_id, "No Value") if gt2_id != "No Value" else "No Value"
        gt3_title = lookup.get(gt3_id, "No Value") if gt3_id != "No Value" else "No Value"

        # Write column F (titles / explanation reasoning)
        f_val = build_f_value(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title)
        ws.cell(row=row, column=6).value = f_val

        # Build prompt and call Claude
        user_msg = build_user_prompt(gt1_id, gt1_title, gt2_id, gt2_title, gt3_id, gt3_title)

        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=400,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}]
            )
            generated_query = response.content[0].text.strip()
        except Exception as e:
            generated_query = f"[ERROR: {e}]"
            print(f"  !! API error at row {row}: {e}")

        # Write column B (query)
        ws.cell(row=row, column=2).value = generated_query

        processed += 1
        pct = int(processed / total * 100)
        print(f"[{row}] {query_id} → GT1={gt1_id} | GT2={gt2_id} | GT3={gt3_id}")
        print(f"       F: {f_val.replace(chr(10), ' | ')}")
        print(f"       B: {generated_query[:100]}{'...' if len(generated_query)>100 else ''}")
        print(f"       Progress: {processed}/{total} ({pct}%)")
        print()

        # Save every 10 rows as checkpoint
        if processed % 10 == 0:
            wb.save(XLSX_PATH)
            print(f"  ── checkpoint saved ({processed} rows done) ──\n")

        time.sleep(API_DELAY)

    # Final save
    wb.save(XLSX_PATH)
    print("=" * 60)
    print(f"Done. Processed: {processed} rows | Skipped: {skipped} rows")
    print(f"Saved to: {XLSX_PATH}")


if __name__ == "__main__":
    main()