import openpyxl
import json
import random

EXCEL_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\QueryDashboard\Sheets\Query_Source_Counter.xlsx'
JSON_PATH = r'D:\Hilmi\Coding\MasterFolderSkripsi\RAG\ragV1\data\iso_controls.json'
SHEET_NAME = 'Explicit Long'
START_ROW = 2
END_ROW = 251

def load_iso_controls(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return {item['control_id']: item for item in data}

def get_control_info(control_id, iso_lookup):
    if control_id is None or control_id == '':
        return None
    return iso_lookup.get(control_id)

INTROS_1 = [
    "Dalam rangka menjamin keamanan sistem informasi di lingkungan organisasi,",
    "Untuk memenuhi standar keamanan informasi yang berlaku,",
    "Dalam konteks pengelolaan keamanan informasi organisasi,",
    "Sebagai bagian dari upaya menjaga integritas dan keamanan sistem informasi,",
    "Dalam kerangka tata kelola keamanan informasi yang baik,",
    "Guna menunjang keamanan dan keandalan sistem informasi organisasi,",
    "Dalam upaya mewujudkan pengelolaan keamanan informasi yang komprehensif,",
    "Sebagai upaya untuk memitigasi risiko keamanan informasi,",
    "Dalam rangka memenuhi persyaratan keamanan informasi organisasi,",
    "Untuk mendukung keberlangsungan operasional sistem informasi yang aman,",
]

INTROS_2 = [
    "Selain itu,",
    "Lebih lanjut,",
    "Di samping itu,",
    "Selanjutnya,",
    "Pada aspek lain,",
    "Dalam hubungannya dengan hal tersebut,",
    "Sejalan dengan hal itu,",
    "Berkaitan dengan aspek tersebut,",
    "Dalam konteks yang lebih luas,",
    "Terkait dengan upaya tersebut,",
]

CONCLUSIONS = [
    "Dengan demikian, penerapan pengendalian tersebut menjadi bagian integral dalam menjaga keamanan dan keberlangsungan layanan sistem informasi organisasi.",
    "Hal ini sejalan dengan kebutuhan organisasi dalam memastikan bahwa seluruh aspek keamanan informasi dikelola secara sistematis dan terdokumentasi dengan baik.",
    "Pengendalian tersebut perlu didukung oleh komitmen seluruh unit kerja serta didokumentasikan dalam bentuk prosedur yang dapat ditinjau secara berkala.",
    "Implementasi pengendalian ini diharapkan mampu mengurangi risiko keamanan informasi serta meningkatkan kesiapan organisasi dalam menghadapi ancaman siber.",
    "Dengan penerapan pengendalian yang konsisten, organisasi dapat memastikan bahwa keamanan informasi terjaga sesuai dengan standar yang ditetapkan.",
    "Oleh karena itu, pengendalian tersebut harus menjadi perhatian utama dalam kerangka manajemen risiko keamanan informasi organisasi.",
    "Penerapan seluruh pengendalian tersebut secara terintegrasi diperlukan guna mencapai tingkat keamanan informasi yang memadai.",
    "Keterpaduan dalam penerapan pengendalian ini menjadi kunci dalam membangun sistem keamanan informasi yang tangguh dan andal.",
]

VERBS = [
    "diperlukan penerapan",
    "organisasi wajib menerapkan",
    "diperlukan penerapan mekanisme",
    "organisasi perlu menyelenggarakan",
    "perlu diimplementasikan",
    "diperlukan pelaksanaan",
    "harus dilaksanakan",
    "organisasi wajib menyelenggarakan",
    "diperlukan pengelolaan",
    "perlu dilakukan",
]

TITLE_TEMPLATES_1 = [
    "{intro} {verb} {title_lower} yang memadai pada seluruh aspek operasional yang terkait.",
    "{intro} {verb} {title_lower} secara konsisten dan berkesinambungan.",
    "{intro} {verb} {title_lower} sesuai dengan ketentuan dan standar yang berlaku.",
]

TITLE_TEMPLATES_2 = [
    "Mekanisme tersebut mencakup {obj_lower} serta pengendalian terhadap potensi risiko yang dapat mengganggu operasional.",
    "Pelaksanaan {title_lower} tersebut harus mencakup langkah-langkah preventif dan detektif yang terstruktur.",
    "Kegiatan ini meliputi {obj_lower} sebagai bagian dari upaya perlindungan menyeluruh.",
    "Pelaksanaannya mencakup serangkaian prosedur yang dirancang untuk meminimalkan risiko keamanan informasi.",
]

TITLE_TEMPLATES_CONNECT = [
    "{intro2} organisasi perlu menyelenggarakan {title2_lower} secara berkelanjutan untuk mendukung keberlangsungan operasional.",
    "{intro2} {title2_lower} harus dilaksanakan secara sistematis sebagai bagian dari kerangka pengendalian internal.",
    "{intro2} pelaksanaan {title2_lower} menjadi esensial dalam menjamin efektivitas pengendalian keamanan informasi.",
    "{intro2} diperlukan penerapan {title2_lower} yang terstruktur dan terdokumentasi dengan baik.",
    "{intro2} organisasi wajib memastikan bahwa {title2_lower} dilakukan secara konsisten dan terukur.",
]

TITLE_TEMPLATES_3 = [
    "Dalam aspek terkait, {title3_lower} juga perlu mendapat perhatian sebagai bagian dari upaya perlindungan menyeluruh.",
    "Selain hal tersebut, pelaksanaan {title3_lower} perlu diintegrasikan ke dalam kerangka pengendalian yang telah ditetapkan.",
    "Pada dimensi lain, {title3_lower} juga merupakan komponen penting yang harus diperhatikan dalam kerangka keamanan informasi.",
]

def build_explanation(gt1, gt2, gt3, iso_lookup):
    parts = []
    if gt1:
        info = get_control_info(gt1, iso_lookup)
        if info:
            parts.append("GT1: " + info['title'])
    if gt2:
        info = get_control_info(gt2, iso_lookup)
        if info:
            parts.append("GT2: " + info['title'])
    elif gt1:
        parts.append("GT2: No Value")
    if gt3:
        info = get_control_info(gt3, iso_lookup)
        if info:
            parts.append("GT3: " + info['title'])
    elif gt2:
        parts.append("GT3: No Value")
    return " | ".join(parts)

def synthesize_long_query(gt1, gt2, gt3, iso_lookup):
    rng = random.Random(hash((gt1, gt2, gt3)))

    controls = []
    for cid in [gt1, gt2, gt3]:
        if cid:
            info = get_control_info(cid, iso_lookup)
            if info:
                controls.append(info)

    if not controls:
        return None

    sentences = []

    intro = rng.choice(INTROS_1)
    verb = rng.choice(VERBS)
    c1 = controls[0]
    title1_lower = c1['title'].lower()
    obj1_lower = c1['objective'].lower().rstrip('.')

    s1 = rng.choice(TITLE_TEMPLATES_1).format(
        intro=intro, verb=verb, title_lower=title1_lower
    )
    sentences.append(s1)

    s2 = rng.choice(TITLE_TEMPLATES_2).format(
        title_lower=title1_lower, obj_lower=obj1_lower
    )
    sentences.append(s2)

    if len(controls) >= 2:
        intro2 = rng.choice(INTROS_2)
        c2 = controls[1]
        title2_lower = c2['title'].lower()
        s3 = rng.choice(TITLE_TEMPLATES_CONNECT).format(
            intro2=intro2, title2_lower=title2_lower
        )
        sentences.append(s3)

    if len(controls) >= 3:
        c3 = controls[2]
        title3_lower = c3['title'].lower()
        s4 = rng.choice(TITLE_TEMPLATES_3).format(title3_lower=title3_lower)
        sentences.append(s4)

    conclusion = rng.choice(CONCLUSIONS)
    sentences.append(conclusion)

    return " ".join(sentences)

def main():
    print("Loading ISO controls...")
    iso_lookup = load_iso_controls(JSON_PATH)
    print(f"Loaded {len(iso_lookup)} controls")

    print("Opening Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    print(f"\nProcessing rows {START_ROW} to {END_ROW} ({END_ROW - START_ROW + 1} queries)...\n")

    processed = 0
    for row in range(START_ROW, END_ROW + 1):
        query_id = ws.cell(row=row, column=1).value
        gt1 = ws.cell(row=row, column=3).value
        gt2 = ws.cell(row=row, column=4).value
        gt3 = ws.cell(row=row, column=5).value

        if not gt1:
            print(f"Row {row}: No GT1, skipping")
            continue

        explanation = build_explanation(gt1, gt2, gt3, iso_lookup)
        query = synthesize_long_query(gt1, gt2, gt3, iso_lookup)

        ws.cell(row=row, column=6).value = explanation
        ws.cell(row=row, column=2).value = query

        processed += 1
        if processed <= 10 or processed % 50 == 0:
            titles_str = ""
            for cid in [gt1, gt2, gt3]:
                if cid:
                    info = get_control_info(cid, iso_lookup)
                    if info:
                        titles_str += info['title'] + " | "
            print(f"{query_id}: GT1={gt1}, GT2={gt2}, GT3={gt3}")
            print(f"  Titles: {titles_str}")
            print(f"  Query: {query[:120]}..." if query and len(query) > 120 else f"  Query: {query}")
            print()

    wb.save(EXCEL_PATH)
    print(f"\nDone! Processed {processed} queries. Saved to {EXCEL_PATH}")

if __name__ == "__main__":
    main()
